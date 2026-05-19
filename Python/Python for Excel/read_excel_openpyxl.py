import pandas as pd
import openpyxl           
import excel
import datetime as dt

# Open the workbook to read cell values.
# The file is automatically closed again after loading the data.
book = openpyxl.load_workbook("xl/stores.xlsx", data_only=True)       

# Get a worksheet object by name or index (0-based)         
sheet = book["2019"]         
sheet = book.worksheets[0]       

# Get a list with all sheet names         
print(book.sheetnames)

# Loop through the sheet objects.
# Instead of "name", openpyxl uses "title".
for i in book.worksheets:             
    print(i.title)

# Getting the dimensions,         
# i.e., the used range of the sheet         
print(sheet.max_row, sheet.max_column)

# Read the value of a single cell         
# using "A1" notation and using cell indices (1-based)
sheet["B6"].value         
print(sheet.cell(row=6, column=2).value)

# Read in a range of cell values by using our excel module         
data = excel.read(book["2019"], (2, 2), (8, 6))         
print(data[:2])  # Print the first two rows