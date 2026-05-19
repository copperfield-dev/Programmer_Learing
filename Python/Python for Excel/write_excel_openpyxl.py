import openpyxl         
from openpyxl.drawing.image import Image         
from openpyxl.chart import BarChart, Reference         
from openpyxl.styles import Font, colors         
from openpyxl.styles.borders import Border, Side         
from openpyxl.styles.alignment import Alignment         
from openpyxl.styles.fills import PatternFill         
import excel
import datetime as dt

# Instantiate a workbook          
book = openpyxl.Workbook()          

# Get the first sheet and give it a name          
sheet = book.active          
sheet.title = "Sheet1"          

# Writing individual cells using A1 notation          
# and cell indices (1-based)          
sheet["A1"].value = "Hello 1"          
sheet.cell(row=2, column=1, value="Hello 2")          

# Formatting: fill color, alignment, border and font          
font_format = Font(color="FF0000", bold=True)          
thin = Side(border_style="thin", color="FF0000")          
sheet["A3"].value = "Hello 3"          
sheet["A3"].font = font_format          
sheet["A3"].border = Border(top=thin, left=thin,                                      
                            right=thin, bottom=thin)          
sheet["A3"].alignment = Alignment(horizontal="center")          
sheet["A3"].fill = PatternFill(fgColor="FFFF00", fill_type="solid")          

# Number formatting (using Excel's formatting strings)          
sheet["A4"].value = 3.3333          
sheet["A4"].number_format = "0.00"          

# Date formatting (using Excel's formatting strings)          
sheet["A5"].value = dt.date(2016, 10, 13)
sheet["A5"].number_format = "mm/dd/yy"

# Formula: you must use the English name of the formula          
# with commas as delimiters          
sheet["A6"].value = "=SUM(A4, 2)"          

# Image          
sheet.add_image(Image("images/python.png"), "C1")          

# Two-dimensional list (we're using our excel module)          
data = [[None, "North", "South"],                  
        ["Last Year", 2, 5],                  
        ["This Year", 3, 6]]          
excel.write(sheet, data, "A10")          

# Chart          
chart = BarChart()          
chart.type = "col"          
chart.title = "Sales Per Region"          
chart.x_axis.title = "Regions"          
chart.y_axis.title = "Sales"          
chart_data = Reference(sheet, min_row=11, min_col=1,                                 
                       max_row=12, max_col=3)          
chart_categories = Reference(sheet, min_row=10, min_col=2,                                      
                             max_row=10, max_col=3)          
# from_rows interprets the data in the same way          
# as if you would add a chart manually in Excel          
chart.add_data(chart_data, titles_from_data=True, from_rows=True)          
chart.set_categories(chart_categories)          
sheet.add_chart(chart, "A15")          

# Saving the workbook creates the file on disk          
book.save("openpyxl.xlsx")